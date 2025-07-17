import os
import json
import pandas as pd
from sqlalchemy import create_engine, text
import sqlalchemy.types
import openpyxl
from datetime import datetime, timedelta

# === CAMINHO DO ARQUIVO EXCEL ===
caminho_excel = r"C:\Users\z223301\Downloads\Software\Cadastros.xlsx"

# === VERIFICA SE O ARQUIVO EXISTE ===
if not os.path.exists(caminho_excel):
    print("❌ Arquivo Excel NÃO encontrado! Verifique o caminho.")
    exit()

print("✅ Arquivo encontrado. Lendo abas...")

# === FUNÇÃO PARA FORMATAR VALORES COMO TEXTO ===
def formatar_valor(cell, col_name):
    """
    Formata valores conforme o tipo de coluna, sempre retornando texto
    """
    if cell.value is None or cell.value == "":
        return ""
    
    valor = cell.value
    
    # Se a coluna contém "Valor", "Pagamento", "Preço", "Custo" - trata como moeda
    if any(palavra in str(col_name).lower() for palavra in ["valor", "pagamento", "preço", "custo", "preco"]):
        try:
            if isinstance(valor, (int, float)):
                return f'R$ {float(valor):,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.')
            elif isinstance(valor, str):
                # Tenta converter string para float
                valor_limpo = valor.replace('R$', '').replace(' ', '').replace('.', '').replace(',', '.')
                return f'R$ {float(valor_limpo):,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.')
        except:
            pass
    
    # Se a coluna contém "Data" ou se o valor é um objeto datetime - trata como data
    if "data" in str(col_name).lower() or isinstance(valor, datetime):
        try:
            if isinstance(valor, datetime):
                return valor.strftime('%d/%m/%Y')
            elif isinstance(valor, (int, float)):
                # Converte número serial do Excel para data
                data_base = datetime(1899, 12, 30)  # Base do Excel
                data_convertida = data_base + timedelta(days=int(valor))
                return data_convertida.strftime('%d/%m/%Y')
            elif isinstance(valor, str):
                # Tenta diferentes formatos de data
                formatos = ['%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y', '%Y/%m/%d']
                for formato in formatos:
                    try:
                        data_obj = datetime.strptime(valor, formato)
                        return data_obj.strftime('%d/%m/%Y')
                    except:
                        continue
        except:
            pass
    
    # Para qualquer outro tipo, retorna como string
    return str(valor)

# === LÊ TODAS AS ABAS DO EXCEL, FORMATANDO COMO TEXTO ===
wb = openpyxl.load_workbook(caminho_excel, data_only=False)
df_abas = {}

for nome_aba in wb.sheetnames:
    print(f"📊 Processando aba: {nome_aba}")
    ws = wb[nome_aba]
    dados = []
    
    # Primeiro, pega os cabeçalhos
    cabecalhos = []
    for j, cell in enumerate(ws[1]):
        cabecalhos.append(str(cell.value) if cell.value is not None else f"Coluna_{j+1}")
    
    # Processa cada linha
    for i, row in enumerate(ws.iter_rows(values_only=False)):
        if i == 0:
            # Linha de cabeçalho
            dados.append(cabecalhos)
        else:
            linha = []
            for j, cell in enumerate(row):
                col_name = cabecalhos[j] if j < len(cabecalhos) else f"Coluna_{j+1}"
                valor_formatado = formatar_valor(cell, col_name)
                linha.append(valor_formatado)
            dados.append(linha)
    
    # Cria DataFrame
    if len(dados) > 1:
        df = pd.DataFrame(dados[1:], columns=dados[0])
        df_abas[nome_aba] = df
        print(f"   ✅ {len(df)} registros processados")
    else:
        print(f"   ⚠️  Aba vazia ou sem dados")

# === CONFIGURAÇÃO DO BANCO NEON (PostgreSQL) ===
usuario = 'neondb_owner'
senha   = 'npg_91HbcvdzrFLw'
host    = 'ep-cold-sky-a537fwxd-pooler.us-east-2.aws.neon.tech'
porta   = '5432'
banco   = 'neondb'

engine = create_engine(
    f'postgresql+psycopg2://{usuario}:{senha}@{host}:{porta}/{banco}'
)

print("\n🔄 Enviando dados para o banco...")

for nome_aba, dados in df_abas.items():
    nome_tabela = nome_aba.strip().lower().replace(" ", "_").replace("-", "_")
    
    try:
        # GARANTE QUE TODOS OS DADOS SEJAM ENVIADOS COMO TEXT
        dados = dados.astype(str)
        
        # Substitui valores 'nan' por string vazia
        dados = dados.replace('nan', '')
        
        # Cria um dicionário de tipos para forçar TEXT em todas as colunas
        tipos_colunas = {col: sqlalchemy.types.TEXT for col in dados.columns}
        
        # 1) Cria/recria a tabela a partir do DataFrame
        dados.to_sql(
            nome_tabela,
            con=engine,
            if_exists='replace',
            index=False,
            dtype=tipos_colunas  # Força todas as colunas como TEXT
        )
        print(f"✅ Tabela '{nome_tabela}' criada/enviada com sucesso.")
        
        # 2) Adiciona a coluna 'id' SERIAL e define como PK
        with engine.begin() as conn:
            # Adiciona a coluna 'id' só se ela não existir
            conn.execute(text(f'''
                DO $$
                BEGIN
                  IF NOT EXISTS (
                    SELECT 1 FROM information_schema.columns
                     WHERE table_name = :tbl
                       AND column_name = 'id'
                  ) THEN
                    ALTER TABLE "{nome_tabela}"
                      ADD COLUMN id SERIAL;
                  END IF;
                END
                $$;
            '''), {"tbl": nome_tabela})
            
            # Define PK sobre id, se ainda não existe
            conn.execute(text(f'''
                DO $$
                BEGIN
                  IF NOT EXISTS (
                    SELECT 1
                      FROM pg_index i
                      JOIN pg_attribute a
                        ON a.attrelid = i.indrelid
                       AND a.attnum = ANY(i.indkey)
                     WHERE i.indrelid = '{nome_tabela}'::regclass
                       AND i.indisprimary
                  ) THEN
                    ALTER TABLE "{nome_tabela}"
                      ADD PRIMARY KEY (id);
                  END IF;
                END
                $$;
            '''))
        print(f"   → Coluna 'id' criada e definida como PK em '{nome_tabela}'.")
        
        # 3) Mostra algumas amostras dos dados formatados
        print(f"   📋 Amostra dos dados formatados:")
        for col in dados.columns[:3]:  # Mostra apenas as 3 primeiras colunas
            if len(dados) > 0:
                valor_exemplo = dados[col].iloc[0] if not dados[col].empty else "N/A"
                print(f"      {col}: {valor_exemplo}")
        
    except Exception as e:
        print(f"❌ Erro na aba '{nome_aba}': {e}")

print("\n🎉 Processamento concluído!")