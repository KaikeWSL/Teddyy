import os
from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QFormLayout,
    QLabel, QLineEdit, QPushButton, QComboBox, QDateEdit,
    QTextEdit, QGroupBox, QRadioButton,
    QFileDialog, QMessageBox, QStackedWidget, QSizePolicy, QTreeView, QAbstractItemView,
    QScrollArea, QProgressDialog, QToolButton, QCalendarWidget, QSpacerItem
)
import shutil
from PyQt6.QtGui import QStandardItemModel, QStandardItem, QPixmap, QIcon
from PyQt6.QtCore import Qt, QDate, QSize, QTimer
from .dialogs import AutocompleteComboBox
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from ..backend.storage_db import load_clientes, load_equipamentos, insert_os, load_tecnicos, insert_tecnico, load_gerentes, insert_gerente, get_os_cadastros, delete_tecnico, delete_gerente
from ..backend.format_utils import format_brl, CurrencyValidator, parse_currency
from .equipamentos_page import EquipamentosPage
from .editar_os_page import EditarOSPage
from .clientes_page import ClientesPage
from .excluir_os_page import ExcluirOSPage
from .abrir_os_page import AbrirOSPage



class CampoDataCustom(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.line = QLineEdit()
        self.line.setPlaceholderText("")
        self.line.setReadOnly(True)
        self.btn = QToolButton()
        self.btn.setText("")
        icon_path = os.path.abspath(os.path.join(
            os.path.dirname(__file__), "..", "imagens", "data.png"
        ))
        self.btn.setIcon(QIcon(icon_path))
        self.btn.setIconSize(QSize(40, 40))
        self.btn.setStyleSheet("""
            QToolButton {
                border: none;
                background: transparent;
                padding: 2px;
            }
            QToolButton:hover {
                background: rgba(0, 0, 0, 0.1);
                border-radius: 2px;
            }
        """)
        self.btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self.btn.setFixedWidth(46)
        self.btn.clicked.connect(self.mostrar_calendario)
        layout = QHBoxLayout(self)
        layout.setContentsMargins(0,0,0,0)
        layout.setSpacing(0)
        layout.addWidget(self.line)
        layout.addWidget(self.btn)
        self.cal = QCalendarWidget()
        self.cal.setWindowFlags(self.cal.windowFlags() | Qt.WindowType.Popup)
        self.cal.clicked.connect(self.selecionar_data)

    def mostrar_calendario(self):
        btn_pos = self.btn.mapToGlobal(self.btn.rect().bottomLeft())
        cal_height = self.cal.sizeHint().height()
        screen = self.cal.screen().geometry()
        # Se não couber para baixo, abre para cima
        if btn_pos.y() + cal_height > screen.bottom():
            pos = self.btn.mapToGlobal(self.btn.rect().topLeft())
            pos.setY(pos.y() - cal_height)
            self.cal.move(pos)
        else:
            self.cal.move(btn_pos)
        self.cal.show()

    def selecionar_data(self, qdate: QDate):
        self.line.setText(qdate.toString("dd/MM/yyyy"))
        self.cal.hide()

    def text(self):
        return self.line.text()

    def setText(self, value):
        self.line.setText(value)

    def clear(self):
        self.line.clear()

    def date(self):
        try:
            return QDate.fromString(self.line.text(), "dd/MM/yyyy")
        except Exception:
            return QDate()

class CadastrarOSPage(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.img_desc_entries = [None] * 6
        self.selected_images = [None] * 6
        self.num_images = 0
        self.clientes = [cli['nome'] for cli in load_clientes()]
        self.equipamentos = [eq['equipamento'] for eq in load_equipamentos()]
        self.entries = {}
        self.status_buttons = []
        self.menu_buttons = []
        self._setup_ui()

    def _setup_ui(self):
        main_layout = QVBoxLayout(self)

        # --- Navegação principal entre os módulos ---
        nav2 = QHBoxLayout()
        for label, page_index in [
            ("Cadastrar OS", 0),
            ("Equipamentos", 1),
            ("Editar OS", 2),
            ("Clientes", 3),
            ("Excluir OS", 4),
            ("Abrir OS", 5),
        ]:
            btn = QPushButton(label)
            btn.setCheckable(True)
            btn.clicked.connect(lambda _, i=page_index: self._on_menu_button_clicked(i))
            btn.setCursor(Qt.CursorShape.PointingHandCursor)
            nav2.addWidget(btn)
            self.menu_buttons.append(btn)
        # Marcar o primeiro como selecionado por padrão
        self.menu_buttons[0].setChecked(True)
        main_layout.addLayout(nav2)

        # --- Cria o stack interno de abas da OS e o QStackedWidget principal ---
        self.stack = QStackedWidget()
        self.content = QStackedWidget()

        # --- Índice 0: container de "Cadastrar OS" com seu próprio nav + stack ---
        cadastrar_widget = QWidget()
        v_cad = QVBoxLayout(cadastrar_widget)

        # nav interno só para "OS", "SERVIÇOS" e "PRÉVIA"
        nav_int = QHBoxLayout()
        self.tab_buttons = []
        for nome, idx in [("Cadastro", 0), ("Serviços", 1), ("Prévia", 2)]:
            btn = QPushButton(nome)
            btn.setCheckable(True)
            btn.setAutoExclusive(True)  
            btn.setObjectName("tabButton")
            btn.clicked.connect(lambda _, i=idx: self._switch_tab(i))
            btn.setCursor(Qt.CursorShape.PointingHandCursor)
            nav_int.addWidget(btn)
            self.tab_buttons.append(btn)
        
        # Seleciona primeira aba por padrão
        self.tab_buttons[0].setChecked(True)
        v_cad.addLayout(nav_int)

        # empacota o próprio stack abaixo dos botões
        v_cad.addWidget(self.stack)
        self.content.addWidget(cadastrar_widget)

        # --- Demais módulos como índices 1–5 ---
        self.content.addWidget(EquipamentosPage(self))
        self.content.addWidget(EditarOSPage(self))
        self.content.addWidget(ClientesPage(self))
        self.content.addWidget(ExcluirOSPage(self))
        self.content.addWidget(AbrirOSPage(self))

        main_layout.addWidget(self.content, 1)

        # --- Popula o stack interno de abas ---
        self.stack.addWidget(self._create_os_tab())
        self.stack.addWidget(self._create_servicos_tab())
        self.stack.addWidget(self._create_previa_tab())

    def _switch_tab(self, index):
        """Muda para a aba especificada"""
        self.stack.setCurrentIndex(index)

    def _create_os_tab(self):
        page = QWidget()
        page_layout = QHBoxLayout(page)
        
        # Valor especial para datas vazias
        data_vazia = QDate(1800, 1, 1)
        # Caminho do ícone de calendário
        icon_data = os.path.abspath(os.path.join(
            os.path.dirname(__file__), "..", "imagens", "data.png"
        ))
        icon_data_url = icon_data.replace("\\", "/")
        dateedit_style = f"""
        QDateEdit::drop-down {{
            subcontrol-origin: padding;
            subcontrol-position: top right;
            width: 36px;
            border: none;
        }}
        QDateEdit::down-arrow {{
            image: url({icon_data_url});
            width: 36px;
            height: 36px;
        }}
        """
        
        altura_fixa = 38  # Altura fixa para todos os campos

        # Grupo de cadastro
        gb3 = QGroupBox("CADASTRO")
        gb3.setObjectName("gbCadastro")
        # Garante altura mínima para não comprimir os campos
        gb3.setMinimumHeight(altura_fixa * 13 + 60)  # 13 campos + margem extra
        gb3.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        form = QFormLayout()
        form.setVerticalSpacing(18)  # Espaçamento vertical fixo
        form.setHorizontalSpacing(8)
        form.setContentsMargins(10, 10, 10, 10)
        gb3.setLayout(form)

        # Configuração de ícones
        icon_path = os.path.abspath(os.path.join(
            os.path.dirname(__file__), "..", "imagens", "seta.png"
        ))
        icon_url = icon_path.replace("\\", "/")

        # OS
        le_os = QLineEdit()
        le_os.setMinimumHeight(altura_fixa)
        le_os.setMaximumHeight(altura_fixa)

        form.addRow("OS:", le_os)
        self.entries["OS"] = le_os

        # Cliente
        cb_cli = AutocompleteComboBox()
        cb_cli.set_completion_list(self.clientes)
        cb_cli.setObjectName("cbCli")
        cb_cli.setMinimumHeight(altura_fixa)
        cb_cli.setMaximumHeight(altura_fixa)
        cb_cli.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        cb_cli.setStyleSheet(f"""
        QComboBox#cbCli {{
            padding-right: 36px;
        }}
        QComboBox#cbCli::drop-down {{
            width: 36px;
            subcontrol-origin: padding;
            subcontrol-position: top right;
            border: none;
        }}
        QComboBox#cbCli::down-arrow {{
            image: url({icon_url});
            width: 32px;
            height: 32px;
        }}
        """)
        form.addRow("Cliente:", cb_cli)
        self.entries["Cliente"] = cb_cli

        # Entrada equip.
        campo_data_in = CampoDataCustom()
        campo_data_in.setMinimumHeight(altura_fixa)
        campo_data_in.setMaximumHeight(altura_fixa)
        campo_data_in.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        campo_data_in.line.setMinimumHeight(altura_fixa)
        campo_data_in.line.setMaximumHeight(altura_fixa)
        campo_data_in.line.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        campo_data_in.btn.setFixedHeight(altura_fixa)
        form.addRow("Entrada equip.:", campo_data_in)
        self.entries["Entrada equip."] = campo_data_in

        # Modelo
        cb_mod = AutocompleteComboBox()
        cb_mod.set_completion_list(self.equipamentos)
        cb_mod.setMinimumHeight(altura_fixa)
        cb_mod.setMaximumHeight(altura_fixa)
        cb_mod.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        cb_mod.setStyleSheet(f"""
        QComboBox {{
            padding-right: 36px;
        }}
        QComboBox::drop-down {{
            width: 36px;
            subcontrol-origin: padding;
            subcontrol-position: top right;
            border: none;
        }}
        QComboBox::down-arrow {{
            image: url({icon_url});
            width: 32px;
            height: 32px;
        }}
        """)
        form.addRow("Modelo:", cb_mod)
        self.entries["Modelo"] = cb_mod

        # N° série
        le_ser = QLineEdit()
        le_ser.setMinimumHeight(altura_fixa)
        le_ser.setMaximumHeight(altura_fixa)
        le_ser.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        form.addRow("N° serie:", le_ser)
        self.entries["N° serie"] = le_ser

        # Valor (currency)
        le_valor = QLineEdit()
        le_valor.setMinimumHeight(altura_fixa)
        le_valor.setMaximumHeight(altura_fixa)
        le_valor.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        le_valor.editingFinished.connect(
            lambda lv=le_valor: lv.setText(format_brl(lv.text()))
        )
        form.addRow("Valor:", le_valor)
        self.entries["Valor"] = le_valor

        # Saída equip.
        campo_data_out = CampoDataCustom()
        campo_data_out.setMinimumHeight(altura_fixa)
        campo_data_out.setMaximumHeight(altura_fixa)
        campo_data_out.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        campo_data_out.line.setMinimumHeight(altura_fixa)
        campo_data_out.line.setMaximumHeight(altura_fixa)
        campo_data_out.line.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        campo_data_out.btn.setFixedHeight(altura_fixa)
        form.addRow("Saída equip.:", campo_data_out)
        self.entries["Saída equip."] = campo_data_out

        # Pagamento (currency)
        le_pag = QLineEdit()
        le_pag.setMinimumHeight(altura_fixa)
        le_pag.setMaximumHeight(altura_fixa)
        le_pag.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        le_pag.editingFinished.connect(
            lambda lp=le_pag: lp.setText(format_brl(lp.text()))
        )
        form.addRow("Pagamento:", le_pag)
        self.entries["Pagamento"] = le_pag

        # Vezes
        le_vezes = QLineEdit()
        le_vezes.setMinimumHeight(altura_fixa)
        le_vezes.setMaximumHeight(altura_fixa)
        le_vezes.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        form.addRow("Vezes:", le_vezes)
        self.entries["Vezes"] = le_vezes

        # Datas de pagamento 1–3
        for i in range(1, 4):
            campo_data = CampoDataCustom()
            campo_data.setMinimumHeight(altura_fixa)
            campo_data.setMaximumHeight(altura_fixa)
            campo_data.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
            campo_data.line.setMinimumHeight(altura_fixa)
            campo_data.line.setMaximumHeight(altura_fixa)
            campo_data.line.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
            campo_data.btn.setFixedHeight(altura_fixa)
            form.addRow(f"Data pagamento {i}:", campo_data)
            self.entries[f"Data pagamento {i}"] = campo_data

        # Adiciona um spacer para garantir espaçamento mesmo em janelas pequenas
        form.addItem(QSpacerItem(20, 40, QSizePolicy.Policy.Minimum, QSizePolicy.Policy.Expanding))

        gb3.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        page_layout.addWidget(gb3, 0)

        # === Avaliação e Causa à direita ===
        right = QVBoxLayout()

        gb1 = QGroupBox("AVALIAÇÃO TÉCNICA")
        te1 = QTextEdit()
        gb1.setLayout(QVBoxLayout())
        gb1.layout().addWidget(te1)
        right.addWidget(gb1)
        self.txt_avaliacao = te1

        gb2 = QGroupBox("CAUSA PROVÁVEL")
        te2 = QTextEdit()
        gb2.setLayout(QVBoxLayout())
        gb2.layout().addWidget(te2)
        right.addWidget(gb2)
        self.txt_causa = te2

        page_layout.addLayout(right, 1)
        return page

    def _create_servicos_tab(self) -> QWidget:
        page = QWidget()
        layout = QHBoxLayout(page)

        # Árvore de serviços
        left = QVBoxLayout()
        self.tree = QTreeView()
        
        # Cria e configura o modelo
        self.model_services = QStandardItemModel()
        self.model_services.setHorizontalHeaderLabels(["Item", "Descrição", "Qtd", "Valor"])
        
        # Associa o modelo à view
        self.tree.setModel(self.model_services)
        self.tree.setEditTriggers(QAbstractItemView.EditTrigger.AllEditTriggers)
        self.model_services.itemChanged.connect(self._on_service_item_changed)
        left.addWidget(self.tree)

        # Botões +/–
        btns = QHBoxLayout()
        b_add = QPushButton("+")
        b_add.clicked.connect(self._add_item)
        b_rem = QPushButton("-")
        b_rem.clicked.connect(self._remove_item)
        b_add.setCursor(Qt.CursorShape.PointingHandCursor)
        b_rem.setCursor(Qt.CursorShape.PointingHandCursor)
        btns.addWidget(b_add)
        btns.addWidget(b_rem)
        left.addLayout(btns)

        # Total & desconto
        hl = QHBoxLayout()
        hl.addWidget(QLabel("TOTAL:"))
        self.le_total = QLineEdit()
        self.le_total.setReadOnly(True)
        hl.addWidget(self.le_total)
        
        self.le_disc = QLineEdit()
        self.le_disc.editingFinished.connect(self._apply_discount)
        self.le_disc.setPlaceholderText("Desconto %")
        hl.addWidget(self.le_disc)
        left.addLayout(hl)

        layout.addLayout(left, 2)

        # Coluna direita
        right = QVBoxLayout()
        icon_path = os.path.abspath(os.path.join(
    os.path.dirname(__file__), "..", "imagens", "seta.png"
))
        icon_url = icon_path.replace("\\", "/")
        # Quantas imagens
        right.addWidget(QLabel("QUANTAS IMAGENS?"))
        self.cb_num = QComboBox()
        self.cb_num.addItems(["0", "1", "3", "6"])
        self.cb_num.setStyleSheet(f"""
        QComboBox#cbCli {{
            padding-right: 36px;
        }}
        QComboBox#cbCli::drop-down {{
            width: 36px;
            subcontrol-origin: padding;
            subcontrol-position: top right;
            border: none;
        }}
        QComboBox#cbCli::down-arrow {{
            image: url({icon_url});
            width: 32px;
            height: 32px;
        }}
        """)
        self.cb_num.currentTextChanged.connect(self._on_num_images)
        right.addWidget(self.cb_num)
        
        # Frames de imagem
        for i in range(6):
            h = QHBoxLayout()
            btn = QPushButton(f"Selecionar imagem {i+1}")
            btn.clicked.connect(lambda _, idx=i: self._select_image(idx))
            btn.setCursor(Qt.CursorShape.PointingHandCursor)
            le = QLineEdit()
            le.setPlaceholderText("Descrição...")
            btn.hide()
            le.hide()
            h.addWidget(btn)
            h.addWidget(le)
            right.addLayout(h)
            self.img_desc_entries[i] = (btn, le)
        
        # Técnico e Gerente (agora ComboBox editáveis)
        right.addWidget(QLabel("TÉCNICO RESPONSÁVEL:"))
        tec_layout = QHBoxLayout()
        tec_layout.setContentsMargins(0, 0, 0, 0)
        tec_layout.setSpacing(4)
        self.cb_tec = QComboBox()
        self.cb_tec.setEditable(True)
        self.cb_tec.setStyleSheet(f"""
        QComboBox#cbCli {{
            padding-right: 36px;
        }}
        QComboBox#cbCli::drop-down {{
            width: 36px;
            subcontrol-origin: padding;
            subcontrol-position: top right;
            border: none;
        }}
        QComboBox#cbCli::down-arrow {{
            image: url({icon_url});
            width: 32px;
            height: 32px;
        }}
        """)
        self.cb_tec.addItems(load_tecnicos())
        self.cb_tec.lineEdit().editingFinished.connect(self._add_tecnico_if_new)
        tec_layout.addWidget(self.cb_tec)
        btn_rem_tec = QPushButton()
        btn_rem_tec.setToolTip("Remove o técnico selecionado da lista e do banco de dados")
        btn_rem_tec.setIcon(QIcon(os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "imagens", "retirar.png"))))
        btn_rem_tec.setIconSize(QSize(16, 16))
        btn_rem_tec.setFixedWidth(32)
        btn_rem_tec.setCursor(Qt.CursorShape.PointingHandCursor)
        btn_rem_tec.clicked.connect(self._remove_tecnico_atual)
        tec_layout.addWidget(btn_rem_tec)
        right.addLayout(tec_layout)
        self.entries["Técnico"] = self.cb_tec
        # Ajusta altura do botão após o layout
        QTimer.singleShot(0, lambda: btn_rem_tec.setFixedHeight(self.cb_tec.height()))

        right.addWidget(QLabel("GERENTE RESPONSÁVEL:"))
        ger_layout = QHBoxLayout()
        ger_layout.setContentsMargins(0, 0, 0, 0)
        ger_layout.setSpacing(4)
        self.cb_ger = QComboBox()
        self.cb_ger.setEditable(True)
        self.cb_ger.setStyleSheet(f"""
        QComboBox#cbCli {{
            padding-right: 36px;
        }}
        QComboBox#cbCli::drop-down {{
            width: 36px;
            subcontrol-origin: padding;
            subcontrol-position: top right;
            border: none;
        }}
        QComboBox#cbCli::down-arrow {{
            image: url({icon_url});
            width: 32px;
            height: 32px;
        }}
        """)
        self.cb_ger.addItems(load_gerentes())
        self.cb_ger.lineEdit().editingFinished.connect(self._add_gerente_if_new)
        ger_layout.addWidget(self.cb_ger)
        btn_rem_ger = QPushButton()
        btn_rem_ger.setToolTip("Remove o gerente selecionado da lista e do banco de dados")
        btn_rem_ger.setIcon(QIcon(os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "imagens", "retirar.png"))))
        btn_rem_ger.setIconSize(QSize(16, 16))
        btn_rem_ger.setFixedWidth(32)
        btn_rem_ger.setCursor(Qt.CursorShape.PointingHandCursor)
        btn_rem_ger.clicked.connect(self._remove_gerente_atual)
        ger_layout.addWidget(btn_rem_ger)
        right.addLayout(ger_layout)
        self.entries["Gerente"] = self.cb_ger
        QTimer.singleShot(0, lambda: btn_rem_ger.setFixedHeight(self.cb_ger.height()))

        # Status do aparelho
        right.addWidget(QLabel("STATUS DO APARELHO:"))
        self.cb_status = QComboBox()
        self.cb_status.setStyleSheet(f"""
        QComboBox#cbCli {{
            padding-right: 36px;
        }}
        QComboBox#cbCli::drop-down {{
            width: 36px;
            subcontrol-origin: padding;
            subcontrol-position: top right;
            border: none;
        }}
        QComboBox#cbCli::down-arrow {{
            image: url({icon_url});
            width: 32px;
            height: 32px;
        }}
        """)
        self.cb_tec.setObjectName("cbCli")
        self.cb_ger.setObjectName("cbCli")
        self.cb_status.setObjectName("cbCli")
        self.cb_num.setObjectName("cbCli")
        self.cb_status.addItems([
            "Orçamento",
            "Aguardando aprovação",
            "Aprovado",
            "Em manutenção",
            "Manutenção em andamento",
            "Manutenção finalizada",
            "Aguardando entrega",
            "Entregue"
        ])
        right.addWidget(self.cb_status)

        self.entries["Status"] = self.cb_status
        right.addStretch()

        layout.addLayout(right, 1)
        return page

    def _create_previa_tab(self) -> QWidget:
        page = QWidget()
        v = QVBoxLayout(page)

        # Área de scroll para a prévia
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)
        
        # Container onde você vai gerar a prévia dinamicamente
        self.prev_widget = QWidget()
        self.prev_container = QVBoxLayout(self.prev_widget)
        scroll.setWidget(self.prev_widget)
        
        v.addWidget(scroll, 1)

        # Botões de ação
        hb = QHBoxLayout()
        btn_preview = QPushButton("Atualizar Prévia")
        btn_preview.clicked.connect(self._preview)
        btn_confirm = QPushButton("Confirmar")
        btn_confirm.clicked.connect(self._on_confirm)
        btn_cancel = QPushButton("Cancelar")
        btn_cancel.clicked.connect(self._on_cancel)
        btn_preview.setCursor(Qt.CursorShape.PointingHandCursor)
        btn_confirm.setCursor(Qt.CursorShape.PointingHandCursor)
        btn_cancel.setCursor(Qt.CursorShape.PointingHandCursor)
        hb.addWidget(btn_preview)
        hb.addWidget(btn_confirm)
        hb.addWidget(btn_cancel)
        v.addLayout(hb)

        return page

    def _on_service_item_changed(self, item: QStandardItem):
        """Sempre que qualquer célula mudar, reformatar se for valor."""
        if item.column() == 3:  # coluna Valor
            try:
                valor = parse_currency(item.text())
                if valor is not None:
                    texto = f"R$ {valor:,.2f}"
                    texto = texto.replace(",", "X").replace(".", ",").replace("X", ".")
                    item.setText(texto)
            except:
                item.setText("R$ 0,00")
        
        self._update_total()

    # — Métodos de imagens (parte direita da aba "SERVIÇOS") —
    def _on_num_images(self, text):
        """Mostra/oculta campos de imagem baseado na quantidade selecionada."""
        try:
            n = int(text)
        except:
            n = 0
        
        self.num_images = n
        for i, (btn, le) in enumerate(self.img_desc_entries):
            visible = i < n
            btn.setVisible(visible)
            le.setVisible(visible)
            if not visible:
                # Limpa dados quando oculta
                self.selected_images[i] = None
                le.clear()

    def _select_image(self, idx):
        """Seleciona arquivo de imagem."""
        path, _ = QFileDialog.getOpenFileName(
            self, "Selecione imagem", 
            filter="Imagens (*.png *.jpg *.jpeg *.bmp *.gif)"
        )
        if not path:
            return
        
        btn, le = self.img_desc_entries[idx]
        btn.setText(f"📷 {os.path.basename(path)}")
        self.selected_images[idx] = path

    def _get_selected_status(self):
        """Retorna o status selecionado."""
        for rb in self.status_buttons:
            if rb.isChecked():
                return rb.text()
        return "Em Orçamento"

    def _os_ja_existe(self, numero_os):
        """Verifica se a OS já existe no banco (case-insensitive)."""
        numero_os = (numero_os or '').strip().lower()
        for registro in get_os_cadastros():
            if str(registro.get('os', '')).strip().lower() == numero_os:
                return True
        return False

    def _preview(self):
        """Gera pré-visualização dos dados inseridos."""
        # Verifica se a OS já existe antes de gerar a prévia
        numero_os = self.entries["OS"].text().strip()
        if self._os_ja_existe(numero_os):
            QMessageBox.critical(self, "Erro", f"Já existe uma OS cadastrada com o número: {numero_os}")
            return
        # Limpa widgets anteriores
        for i in reversed(range(self.prev_container.count())):
            w = self.prev_container.itemAt(i).widget()
            if w:
                w.setParent(None)

        # Título
        title = QLabel("📋 PRÉVIA DA ORDEM DE SERVIÇO")
        title.setStyleSheet("font-size: 16px; font-weight: bold; margin: 10px;")
        self.prev_container.addWidget(title)

        # Dados básicos
        dados_group = QGroupBox("Dados Básicos")
        dados_layout = QFormLayout(dados_group)

        dados_layout.addRow("OS:", QLabel(self.entries["OS"].text() or "Não informado"))
        dados_layout.addRow("Cliente:", QLabel(self.entries["Cliente"].currentText() or "Não selecionado"))
        dados_layout.addRow("Modelo:", QLabel(self.entries["Modelo"].currentText() or "Não selecionado"))
        dados_layout.addRow("N° Série:", QLabel(self.entries["N° serie"].text() or "Não informado"))
        dados_layout.addRow("Entrada:", QLabel(self.entries["Entrada equip."].text() or "Não informado"))
        dados_layout.addRow("Saída:", QLabel(self.entries["Saída equip."].text() or "Não informado"))
        dados_layout.addRow("Valor:", QLabel(self.entries["Valor"].text() or "R$ 0,00"))
        dados_layout.addRow("Pagamento:", QLabel(self.entries["Pagamento"].text() or "R$ 0,00"))
        dados_layout.addRow("Vezes:", QLabel(self.entries["Vezes"].text() or "1"))

        self.prev_container.addWidget(dados_group)

        # Serviços
        servicos_group = QGroupBox("Serviços")
        servicos_layout = QVBoxLayout(servicos_group)

        if self.model_services.rowCount() == 0:
            servicos_layout.addWidget(QLabel("Nenhum serviço adicionado"))
        else:
            for row in range(self.model_services.rowCount()):
                item = self.model_services.item(row, 0).text()
                desc = self.model_services.item(row, 1).text()
                qtd = self.model_services.item(row, 2).text()
                valor = self.model_services.item(row, 3).text()

                servico_text = f"{item}. {desc} - Qtd: {qtd} - Valor: {valor}"
                servicos_layout.addWidget(QLabel(servico_text))

            total_label = QLabel(f"Total: {self.le_total.text()}")
            total_label.setStyleSheet("font-weight: bold;")
            servicos_layout.addWidget(total_label)

        self.prev_container.addWidget(servicos_group)

        # Avaliação e Causa
        avaliacao_group = QGroupBox("Avaliação Técnica")
        avaliacao_layout = QVBoxLayout(avaliacao_group)
        avaliacao_text = QLabel(self.txt_avaliacao.toPlainText() or "Não informado")
        avaliacao_text.setWordWrap(True)
        avaliacao_layout.addWidget(avaliacao_text)
        self.prev_container.addWidget(avaliacao_group)

        causa_group = QGroupBox("Causa Provável")
        causa_layout = QVBoxLayout(causa_group)
        causa_text = QLabel(self.txt_causa.toPlainText() or "Não informado")
        causa_text.setWordWrap(True)
        causa_layout.addWidget(causa_text)
        self.prev_container.addWidget(causa_group)

        # === Seção de Imagens Atualizada ===
        if self.num_images > 0:
            imagem_group = QGroupBox(f"Imagens ({self.num_images})")
            imagem_layout = QVBoxLayout(imagem_group)

            for i in range(self.num_images):
                container_h = QHBoxLayout()

                if self.selected_images[i]:
                    # Carrega o QPixmap e redimensiona para thumbnail
                    pix = QPixmap(self.selected_images[i])
                    # Ajuste a largura máxima ou altura máxima conforme quiser:
                    pix = pix.scaledToWidth(100, Qt.TransformationMode.SmoothTransformation)

                    lbl_img = QLabel()
                    lbl_img.setPixmap(pix)
                    lbl_img.setFixedSize(pix.size())  # Garante que o QLabel tenha o tamanho exato do thumbnail

                    # Texto com descrição ao lado da imagem
                    descricao = self.img_desc_entries[i][1].text().strip() or "Sem descrição"
                    lbl_desc = QLabel(f"{os.path.basename(self.selected_images[i])} – {descricao}")
                    lbl_desc.setWordWrap(True)

                    container_h.addWidget(lbl_img, 0, Qt.AlignmentFlag.AlignLeft)
                    container_h.addWidget(lbl_desc, 1, Qt.AlignmentFlag.AlignVCenter)
                else:
                    # Caso não tenha imagem selecionada, avisa o usuário
                    lbl_sem = QLabel(f"{i+1}. Imagem não selecionada")
                    container_h.addWidget(lbl_sem)

                imagem_layout.addLayout(container_h)

            self.prev_container.addWidget(imagem_group)

        # Status e responsáveis
        status_group = QGroupBox("Status e Responsáveis")
        status_layout = QFormLayout(status_group)

        status_layout.addRow("Status:", QLabel(self.cb_status.currentText() or "Não informado"))
        status_layout.addRow("Técnico:", QLabel(self.cb_tec.currentText() or "Não informado"))
        status_layout.addRow("Gerente:", QLabel(self.cb_ger.currentText() or "Não informado"))

        self.prev_container.addWidget(status_group)

    def _validate_data(self):
        """Valida os dados antes do cadastro."""
        errors = []
        
        if not self.entries["OS"].text().strip():
            errors.append("- Número da OS é obrigatório")
        
        if not self.entries["Cliente"].currentText():
            errors.append("- Cliente deve ser selecionado")
        
        if not self.entries["Modelo"].currentText():
            errors.append("- Modelo deve ser selecionado")
        
        if not self.entries["N° serie"].text().strip():
            errors.append("- Número de série é obrigatório")
        
        if not self.entries["Técnico"].text().strip():
            errors.append("- Técnico responsável é obrigatório")
        
        # Verifica se há serviços
        if self.model_services.rowCount() == 0:
            errors.append("- Pelo menos um serviço deve ser adicionado")
        else:
            # Valida serviços
            for row in range(self.model_services.rowCount()):
                desc_item = self.model_services.item(row, 1)
                if not desc_item or not desc_item.text().strip():
                    errors.append(f"- Serviço {row+1}: Descrição é obrigatória")
        
        # Verifica imagens selecionadas
        for i in range(self.num_images):
            if not self.selected_images[i]:
                errors.append(f"- Imagem {i+1} não foi selecionada")
        
        return errors

    def _insert_images_in_excel(self, workbook_path: str, sheet_name: str):
        """
        Insere as imagens selecionadas na planilha Excel na aba `sheet_name`.
        No seu template (Orçamento.xlsx) você deve ter placeholders nomeados:
            IMAGEM_1, IMAGEM_2, ..., IMAGEM_6
        Como formas (Shapes) na aba correta.
        Se não encontrar a forma, joga a imagem em uma célula fallback (A10+i).
        """
        if not os.path.exists(workbook_path):
            raise FileNotFoundError(f"Arquivo Excel não encontrado: {workbook_path}")

        wb = load_workbook(workbook_path)
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Aba '{sheet_name}' não existe em {workbook_path}")

        ws = wb[sheet_name]

        for i in range(self.num_images):
            img_path = self.selected_images[i]
            if img_path and os.path.exists(img_path):
                shape_name = f"IMAGEM_{i+1}"
                img = OpenpyxlImage(img_path)

                shape_found = False
                # Percorre todas as shapes da planilha para achar a com name == IMAGEM_{i+1}
                for shape in ws._shapes:
                    if hasattr(shape, "name") and shape.name == shape_name:
                        # Ajusta tamanho da imagem conforme a forma
                        if hasattr(shape, "width"):
                            img.width = shape.width
                        if hasattr(shape, "height"):
                            img.height = shape.height
                        # Define o mesmo âncora (posição) da forma
                        if hasattr(shape, "anchor"):
                            img.anchor = shape.anchor
                        ws.add_image(img)
                        shape_found = True
                        break

                if not shape_found:
                    # Se não existirem shapes nomeadas, joga em célula fallback:
                    cell_ref = f"A{10 + i}"
                    img.anchor = cell_ref
                    ws.add_image(img)

        wb.save(workbook_path)
        wb.close()

    def _on_confirm(self):
        """
        1) Valida dados do formulário
        2) Carrega clientes e equipamentos APENAS UMA VEZ
        3) Copia o template 'Orçamento.xlsx' para 'Z:\\OS\\<Cliente>\\<OS>\\orçamento.xlsx'
        4) Preenche a aba correta e exclui as abas não usadas
        5) Salva o orçamento completo
        6) Gera 'orçamento_tecnicos.xlsx' a partir desse arquivo, mantendo só a aba usada
        e truncando tudo abaixo da linha 52 (colunas A-I)
        7) Exporta ambos para PDF
        8) Insere no banco e exibe sucesso
        """
        import os, shutil
        from openpyxl import load_workbook
        from openpyxl.drawing.image import Image as OpenpyxlImage
        from openpyxl.utils.cell import (
            coordinate_from_string,
            column_index_from_string,
            get_column_letter
        )

        # --- Função auxiliar para formatação de data ---
        def format_date_to_standard(date_obj):
            """
            Converte QDate para formato padrão 'dd/MM/yyyy'
            Verifica se a data é válida antes de converter
            """
            if date_obj is None:
                return ""
            
            # Verifica se a data é válida (não é nula)
            if not date_obj.isValid():
                return ""
                
            return date_obj.toString('dd/MM/yyyy')
        
        # --- Função auxiliar para formatação de moeda ---
        def format_currency_to_brazilian(value):
            """
            Converte valor numérico para formato de moeda brasileira
            Ex: 1500.50 -> "R$ 1.500,50"
            """
            if value is None or value == 0:
                return "R$ 0,00"
            
            # Converte para string com 2 casas decimais
            valor_str = f"{value:.2f}"
            
            # Separa parte inteira e decimal
            partes = valor_str.split('.')
            parte_inteira = partes[0]
            parte_decimal = partes[1]
            
            # Adiciona pontos como separadores de milhares
            parte_inteira_formatada = ""
            for i, digito in enumerate(reversed(parte_inteira)):
                if i > 0 and i % 3 == 0:
                    parte_inteira_formatada = "." + parte_inteira_formatada
                parte_inteira_formatada = digito + parte_inteira_formatada
            
            return f"R$ {parte_inteira_formatada},{parte_decimal}"

        # --- 1) Isolar dados do formulário e validar ---
        try:
            num_images = object.__getattribute__(self, 'num_images')
        except AttributeError:
            num_images = 0

        # Debug: Verificar se as datas estão sendo capturadas corretamente
        entrada_date = self.entries['Entrada equip.'].date()
        saida_date = self.entries['Saída equip.'].date()
        
        print(f"Debug - Entrada date: {entrada_date.toString('yyyy-MM-dd')}")
        print(f"Debug - Saída date: {saida_date.toString('yyyy-MM-dd')}")

        data_dict = {
            'os':                self.entries['OS'].text().strip(),
            'cliente':           self.entries['Cliente'].currentText(),
            'modelo':            self.entries['Modelo'].currentText(),
            'entrada_equip':     format_date_to_standard(entrada_date),
            'valor':             format_currency_to_brazilian(parse_currency(self.entries['Valor'].text())),
            'saida_equip':       format_date_to_standard(saida_date),
            'pagamento':         format_currency_to_brazilian(parse_currency(self.entries['Pagamento'].text())),
            'vezes':             int(self.entries['Vezes'].text() or 0),
            'data_pag1':         format_date_to_standard(self.entries['Data pagamento 1'].date()),
            'data_pag2':         format_date_to_standard(self.entries['Data pagamento 2'].date()),
            'data_pag3':         format_date_to_standard(self.entries['Data pagamento 3'].date()),
            'serie':             self.entries['N° serie'].text(),
            'tecnico':           self.cb_tec.currentText().strip(),
            'avaliacao_tecnica': self.txt_avaliacao.toPlainText(),
            'causa_provavel':    self.txt_causa.toPlainText(),
            'status':            self.cb_status.currentText(),
        }
        
        # Debug: Verificar os dados formatados
        print(f"Debug - entrada_equip formatada: '{data_dict['entrada_equip']}'")
        print(f"Debug - saida_equip formatada: '{data_dict['saida_equip']}'")
        
        # Validações adicionais para datas
        if not data_dict['entrada_equip']:
            erros.append("• Data de entrada do equipamento é obrigatória.")


        erros = []
        if not data_dict['os']:
            erros.append("• Número da OS é obrigatório.")
        if not data_dict['cliente']:
            erros.append("• Cliente não foi escolhido.")
        if not data_dict['modelo']:
            erros.append("• Modelo não foi escolhido.")
        if parse_currency(self.entries['Valor'].text()) is None:
            erros.append("• Valor inválido ou não informado.")
        if not data_dict['serie']:
            erros.append("• N° Série é obrigatório.")


        # Verifica se há serviços
        if self.model_services.rowCount() == 0:
            erros.append("• Pelo menos um serviço deve ser adicionado")
        else:
            # Valida serviços
            for row in range(self.model_services.rowCount()):
                desc_item = self.model_services.item(row, 1)
                if not desc_item or not desc_item.text().strip():
                    erros.append(f"• Serviço {row+1}: Descrição é obrigatória")

        # Verifica imagens selecionadas
        for i in range(num_images):
            if not self.selected_images[i]:
                erros.append(f"• Imagem {i+1} não foi selecionada")

        if erros:
            QMessageBox.critical(None, "Erro de Validação", "\n".join(erros))
            return

        # Verifica se a OS já existe antes de cadastrar
        numero_os = self.entries["OS"].text().strip()
        if self._os_ja_existe(numero_os):
            QMessageBox.critical(self, "Erro", f"Já existe uma OS cadastrada com o número: {numero_os}")
            return

        # --- 2) Carregar CLIENTES e EQUIPAMENTOS apenas uma vez ---
        try:
            clientes_list = load_clientes()         # lista de dicts
            equipamentos_list = load_equipamentos() # lista de dicts
        except Exception as e:
            QMessageBox.critical(None, "Erro",
                                f"Falha ao carregar dados do banco:\n{e}")
            return

        # --- Função auxiliar para merged cells (mesma que você já tinha) ---
        def _write_merged(ws, coord, valor):
            """
            Se coord (ex.: "B15") fizer parte de merged cells,
            escreve em seu canto superior‐esquerdo (anchor).
            Caso contrário, escreve em coord normalmente.
            """
            col_letra, row = coordinate_from_string(coord)
            col_idx = column_index_from_string(col_letra)
            for merged in ws.merged_cells.ranges:
                if (merged.min_row <= row <= merged.max_row
                        and merged.min_col <= col_idx <= merged.max_col):
                    # achou intervalo mesclado que contém coord
                    col_letra = get_column_letter(merged.min_col)
                    row = merged.min_row
                    coord = f"{col_letra}{row}"
                    break
            ws[coord] = valor

        # --- 3) Preparar pastas em Z:\OS\<Cliente>\<OS> (com sanitização) ---
        raiz_os = r"Z:\OS"

        def _sanitize_folder_name(name: str) -> str:
            """
            Retira espaços no início/fim, substitui caracteres proibidos no Windows
            (\ / : * ? " < > |) por '_' e remove pontos/espaços finais.
            """
            s = name.strip()
            for ch in ['\\', '/', ':', '*', '?', '"', '<', '>', '|']:
                s = s.replace(ch, '_')
            s = s.rstrip('. ')
            return s

        cliente_raw = data_dict['cliente']
        os_raw      = data_dict['os']

        cliente_seguro = _sanitize_folder_name(cliente_raw)
        os_seguro      = _sanitize_folder_name(os_raw)

        cliente_dir = os.path.join(raiz_os, cliente_seguro)
        os_dir      = os.path.join(cliente_dir, os_seguro)

        try:
            os.makedirs(os_dir, exist_ok=True)
        except Exception as e:
            QMessageBox.critical(None, "Erro",
                                f"Não foi possível criar pasta:\n{e}")
            return

        # --- 3.2) Definir caminhos completos para os arquivos de saída ---
        full_path_orcamento = os.path.join(os_dir, "orçamento.xlsx")
        full_path_tecnicos  = os.path.join(os_dir, "orçamento_tecnicos.xlsx")
        full_path_orc_pdf   = os.path.join(os_dir, "orçamento.pdf")
        full_path_tec_pdf   = os.path.join(os_dir, "orçamento_tecnicos.pdf")

        # --- 4) Copiar template para orçamento.xlsx e preencher abas ---
        # template_path = os.path.join(os.getcwd(), "Orçamento.xlsx")
        # Corrigido para buscar na raiz do projeto
        import sys
        base_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
        template_path = os.path.join(base_dir, "Orçamento.xlsx")
        if not os.path.exists(template_path):
            QMessageBox.critical(None, "Erro",
                                f"Template não encontrado:\n{template_path}")
            return

        try:
            shutil.copyfile(template_path, full_path_orcamento)
        except Exception as e:
            QMessageBox.critical(None, "Erro",
                                f"Falha ao copiar template:\n{e}")
            return

        try:
            wb_dest = load_workbook(full_path_orcamento)
        except Exception as e:
            QMessageBox.critical(None, "Erro",
                                f"Falha ao abrir cópia do template:\n{e}")
            return

        # Determinar qual aba conterá as imagens
        sheet_map  = {0: 'SEM_IMAGEM', 1: '1_IMAGEM', 3: '3_IMAGEM', 6: '6_IMAGEM'}
        aba_imagem = sheet_map.get(num_images, 'SEM_IMAGEM')
        if aba_imagem not in wb_dest.sheetnames:
            QMessageBox.critical(None, "Erro",
                                f"Aba '{aba_imagem}' não existe em Orçamento.xlsx.")
            wb_dest.close()
            return

        # 4.1) Excluir todas as abas exceto a usada
        for aba in list(wb_dest.sheetnames):
            if aba != aba_imagem:
                del wb_dest[aba]

        # 4.2) Mapear coordenadas fixas para miniaturas, conforme layout do seu template
        image_cell_map = {
            '1_IMAGEM': ["B31"],
            '3_IMAGEM': ["A37", "E37", "I37"],
            '6_IMAGEM': ["A30", "E30", "I30", "A41", "E41", "I41"]
        }

        # 4.3) Converter cm → pixels (para 96 DPI)
        px_por_cm = 96.0 / 2.54

        # 4.4) Preencher a aba restante
        ws = wb_dest[aba_imagem]

        # Campos fixos:
        _write_merged(ws, 'I3',  data_dict['os'])
        _write_merged(ws, 'B12', data_dict['cliente'])

        # Dados do cliente a partir da lista carregada
        for cli in clientes_list:
            if cli['nome'] == data_dict['cliente']:
                _write_merged(ws, 'B14', cli.get('cpf_cnpj')    or '')
                _write_merged(ws, 'B13', cli.get('email')       or '')
                _write_merged(ws, 'H12', cli.get('endereco')    or '')
                _write_merged(ws, 'H14', cli.get('tel_contato') or '')
                _write_merged(ws, 'B15', cli.get('tel_celular') or '')
                break

        # Avaliação, causa e valor
        _write_merged(ws, 'A23', data_dict['avaliacao_tecnica'])
        _write_merged(ws, 'A26', data_dict['causa_provavel'])
        _write_merged(ws, 'I73', data_dict['valor'])

        # Linhas de serviços (linhas 60 a 72)
        for idx in range(self.model_services.rowCount()):
            row = 60 + idx
            ws[f'A{row}'] = self.model_services.item(idx, 0).text()
            ws[f'B{row}'] = self.model_services.item(idx, 1).text()
            ws[f'H{row}'] = self.model_services.item(idx, 2).text()
            ws[f'I{row}'] = self.model_services.item(idx, 3).text()

        # Texto descritivo em A18
        equip = data_dict['modelo']
        tipo  = next((e['tipo']  for e in equipamentos_list if e['equipamento'] == equip), '')
        marca = next((e['marca'] for e in equipamentos_list if e['equipamento'] == equip), '')
        _write_merged(
            ws,
            'A18',
            f'Aparelho "{tipo}", marca "{marca}", modelo "{equip}", '
            f'com número de série "{data_dict["serie"]}", sem acessórios e com ordem: "{data_dict["os"]}"'
        )

        # Inserir miniaturas na aba de imagem
        if num_images > 0:
            coords = image_cell_map.get(aba_imagem, [])
            for i in range(num_images):
                if i >= len(coords):
                    break
                img_path = self.selected_images[i]
                if not img_path or not os.path.exists(img_path):
                    continue

                cell_coord = coords[i]
                img_obj = OpenpyxlImage(img_path)

                if aba_imagem == '1_IMAGEM':
                    # 10×15 cm
                    img_obj.height = int(px_por_cm * 10)  # ~378 px
                    img_obj.width  = int(px_por_cm * 15)  # ~567 px
                else:
                    # 4×6 cm
                    img_obj.height = int(px_por_cm * 4)   # ~151 px
                    img_obj.width  = int(px_por_cm * 6)   # ~227 px

                img_obj.anchor = cell_coord
                ws.add_image(img_obj)

        # --- 5) Salvar orçamento.xlsx (já com só a aba usada) ---
        try:
            wb_dest.save(full_path_orcamento)
            wb_dest.close()
        except Exception as e:
            QMessageBox.critical(None, "Erro",
                                f"Falha ao salvar orçamento completo:\n{e}")
            return

        # --- 6) Gerar orçamento_tecnicos.xlsx a partir do orçamento completo ---
        try:
            shutil.copyfile(full_path_orcamento, full_path_tecnicos)
        except Exception as e:
            QMessageBox.critical(None, "Erro",
                                f"Falha ao copiar para versão técnicos:\n{e}")
            return

        # 6.1) Abrir orçamento_tecnicos.xlsx, truncar abaixo da linha 52 (apenas na aba de imagem)
        try:
            wb_tec = load_workbook(full_path_tecnicos)
        except Exception as e:
            QMessageBox.critical(None, "Erro",
                                f"Falha ao abrir orçamento_tecnicos.xlsx:\n{e}")
            return

        if aba_imagem in wb_tec.sheetnames:
            ws_tec = wb_tec[aba_imagem]
            # Excluir todas as linhas abaixo da 52
            try:
                while ws_tec.max_row >= 53:
                    ws_tec.delete_rows(53)
            except Exception:
                # Se não conseguir apagar todas de uma vez, ignora
                pass

        # Salvar a versão técnicos
        try:
            wb_tec.save(full_path_tecnicos)
            wb_tec.close()
        except Exception as e:
            QMessageBox.critical(None, "Erro",
                                f"Falha ao salvar orçamento_tecnicos.xlsx:\n{e}")
            return

        # --- 7) Exportar ambos para PDF (se win32com disponível) ---
        try:
            import win32com.client as win32
            excel_app = win32.Dispatch('Excel.Application')
            excel_app.Visible = False
            excel_app.DisplayAlerts = False

            # Primeiro, orçamento.xlsx → orçamento.pdf
            wb_pdf = excel_app.Workbooks.Open(full_path_orcamento)
            wb_pdf.ExportAsFixedFormat(0, full_path_orc_pdf)
            wb_pdf.Close(False)

            # Depois, orçamento_tecnicos.xlsx → orçamento_tecnicos.pdf
            wb_pdf2 = excel_app.Workbooks.Open(full_path_tecnicos)
            wb_pdf2.ExportAsFixedFormat(0, full_path_tec_pdf)
            wb_pdf2.Close(False)

            excel_app.Quit()

            # Verifica se os PDFs foram criados
            if not os.path.isfile(full_path_orc_pdf):
                QMessageBox.warning(None, "Aviso", f"O PDF do orçamento não foi criado em:\n{full_path_orc_pdf}")
            if not os.path.isfile(full_path_tec_pdf):
                QMessageBox.warning(None, "Aviso", f"O PDF do orçamento de técnicos não foi criado em:\n{full_path_tec_pdf}")
        except Exception as e:
            # Se não tiver win32com ou der erro, apenas ignora sem interromper
            QMessageBox.warning(None, "Aviso", f"Erro ao exportar para PDF:\n{e}")

        # --- 8) Inserir no banco e exibe sucesso ---
        data_dict['id'] = None  # Garante que o argumento 'id' seja passado
        try:
            insert_os(**data_dict)
        except Exception as e:
            QMessageBox.critical(None, "Erro",
                                f"Falha ao gravar no banco:\n{e}")
            return

        QMessageBox.information(None, 'Sucesso',
                                'OS cadastrada com sucesso.')
        self._on_cancel()

    def _on_cancel(self):
        """
        Limpa todos os campos do formulário e da pré-visualização.

        Agora protege cada widget.clear() contra RuntimeError, caso o widget já tenha sido deletado.
        """
        # 1) Limpar QLineEdits e QComboBoxes em self.entries
        from PyQt6.QtWidgets import QLineEdit, QComboBox
        for key, widget in list(self.entries.items()):
            try:
                if widget is not None:
                    if isinstance(widget, QLineEdit):
                        widget.clear()
                    elif isinstance(widget, QComboBox):
                        widget.setCurrentIndex(0)
                    else:
                        widget.clear()
            except RuntimeError:
                pass

        # 2) Limpar QTextEdits
        try:
            if hasattr(self, 'txt_avaliacao') and self.txt_avaliacao:
                self.txt_avaliacao.clear()
        except RuntimeError:
            pass

        try:
            if hasattr(self, 'txt_causa') and self.txt_causa:
                self.txt_causa.clear()
        except RuntimeError:
            pass

        # 3) Limpar dados de serviços (árvore)
        try:
            if hasattr(self, 'model_services') and self.model_services:
                self.model_services.removeRows(0, self.model_services.rowCount())
        except RuntimeError:
            pass

        # 4) Resetar combo de número de imagens e esconder botões de imagem
        try:
            if hasattr(self, 'cb_num') and self.cb_num:
                self.cb_num.setCurrentIndex(0)
        except RuntimeError:
            pass

        for img_tuple in getattr(self, 'img_desc_entries', []) or []:
            try:
                btn, le = img_tuple
                if btn:
                    btn.hide()
                if le:
                    le.hide()
                    le.clear()
            except RuntimeError:
                pass

        # 5) Limpar QLineEdits de técnico e gerente
        try:
            if hasattr(self, 'cb_tec') and self.cb_tec:
                self.cb_tec.setCurrentIndex(0)
        except RuntimeError:
            pass

        try:
            if hasattr(self, 'cb_ger') and self.cb_ger:
                self.cb_ger.setCurrentIndex(0)
        except RuntimeError:
            pass

        # 6) Limpar pré-visualização (widgets gerados dinamicamente)
        try:
            # Se existir prev_container (VBoxLayout), remova todos os widgets filhos
            if hasattr(self, 'prev_container') and self.prev_container:
                for i in reversed(range(self.prev_container.count())):
                    child = self.prev_container.itemAt(i).widget()
                    if child:
                        child.setParent(None)
        except RuntimeError:
            pass

        # 7) Restaurar botões de status (radio buttons) para o estado inicial
        try:
            for rb in getattr(self, 'status_buttons', []):
                if rb:
                    rb.setChecked(False)
            # Marcar o primeiro (em orçamento) por padrão
            if self.status_buttons:
                self.status_buttons[0].setChecked(True)
        except RuntimeError:
            pass

        # 8) Resetar total e desconto
        try:
            if hasattr(self, 'le_total') and self.le_total:
                self.le_total.clear()
        except RuntimeError:
            pass

        try:
            if hasattr(self, 'le_disc') and self.le_disc:
                self.le_disc.clear()
        except RuntimeError:
            pass

        # 9) Resetar datas para hoje
        try:
            if hasattr(self, 'entries') and 'Entrada equip.' in self.entries:
                self.entries['Entrada equip.'].setText(QDate.currentDate().toString("dd/MM/yyyy"))
        except RuntimeError:
            pass

        try:
            if hasattr(self, 'entries') and 'Saída equip.' in self.entries:
                self.entries['Saída equip.'].setText(QDate.currentDate().toString("dd/MM/yyyy"))
        except RuntimeError:
            pass

        for i in range(1, 4):
            key = f"Data pagamento {i}"
            try:
                if hasattr(self, 'entries') and key in self.entries:
                    self.entries[key].setText(QDate.currentDate().toString("dd/MM/yyyy"))
            except RuntimeError:
                pass

        # 10) Resetar combo de garantia


    def _add_tecnico_if_new(self):
        nome = self.cb_tec.currentText().strip()
        if nome and nome not in [self.cb_tec.itemText(i) for i in range(self.cb_tec.count())]:
            insert_tecnico(nome)
            self.cb_tec.addItem(nome)

    def _add_gerente_if_new(self):
        nome = self.cb_ger.currentText().strip()
        if nome and nome not in [self.cb_ger.itemText(i) for i in range(self.cb_ger.count())]:
            insert_gerente(nome)
            self.cb_ger.addItem(nome)

    def _on_menu_button_clicked(self, page_index):
        self.content.setCurrentIndex(page_index)
        for i, btn in enumerate(self.menu_buttons):
            btn.setChecked(i == page_index)

    def _add_item(self):
        """Adiciona uma nova linha numerada e já com Valor zerado em R$."""
        idx = self.model_services.rowCount() + 1
        from PyQt6.QtGui import QStandardItem
        items = [
            QStandardItem(str(idx)),
            QStandardItem(""),         # Descrição
            QStandardItem("1"),        # Quantidade padrão
            QStandardItem("R$ 0,00")   # Valor inicial
        ]
        for it in items:
            it.setEditable(True)
        self.model_services.appendRow(items)
        self._update_total()

    def _remove_item(self):
        """Remove as linhas selecionadas e renumera a coluna Item."""
        selection = self.tree.selectionModel().selectedRows()
        if not selection:
            from PyQt6.QtWidgets import QMessageBox
            QMessageBox.information(self, "Aviso", "Selecione uma linha para remover.")
            return
        for index in sorted(selection, key=lambda idx: idx.row(), reverse=True):
            self.model_services.removeRow(index.row())
        for row in range(self.model_services.rowCount()):
            self.model_services.item(row, 0).setText(str(row + 1))
        self._update_total()

    def _apply_discount(self):
        """Aplica desconto percentual ao total."""
        try:
            pct = float(self.le_disc.text().replace(",", "."))
            if pct < 0 or pct > 100:
                raise ValueError("Desconto deve estar entre 0% e 100%")
        except ValueError as e:
            from PyQt6.QtWidgets import QMessageBox
            QMessageBox.critical(self, "Erro", f"Desconto inválido: {e}")
            self.le_disc.clear()
            return
        tot = 0.0
        for row in range(self.model_services.rowCount()):
            valor_item = self.model_services.item(row, 3)
            if valor_item:
                valor = parse_currency(valor_item.text())
                if valor:
                    tot += valor
        disc = tot * (1 - pct/100)
        txt = f"R$ {disc:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
        self.le_total.setText(txt)

    def _update_total(self):
        """Atualiza o total baseado nos valores dos serviços."""
        tot = 0.0
        for row in range(self.model_services.rowCount()):
            valor_item = self.model_services.item(row, 3)
            if valor_item:
                text = valor_item.text()
                valor = parse_currency(text)
                if valor:
                    tot += valor
        txt = f"R$ {tot:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
        self.le_total.setText(txt)

    def _remove_tecnico_atual(self):
        idx = self.cb_tec.currentIndex()
        nome = self.cb_tec.currentText().strip()
        if idx > 0 and nome:
            reply = QMessageBox.question(self, "Remover Técnico", f"Deseja realmente remover o técnico '{nome}' do banco de dados?", QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
            if reply == QMessageBox.StandardButton.Yes:
                try:
                    delete_tecnico(nome)
                except Exception as e:
                    QMessageBox.critical(self, "Erro", f"Erro ao remover técnico do banco:\n{e}")
                    return
                self.cb_tec.removeItem(idx)

    def _remove_gerente_atual(self):
        idx = self.cb_ger.currentIndex()
        nome = self.cb_ger.currentText().strip()
        if idx > 0 and nome:
            reply = QMessageBox.question(self, "Remover Gerente", f"Deseja realmente remover o gerente '{nome}' do banco de dados?", QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
            if reply == QMessageBox.StandardButton.Yes:
                try:
                    delete_gerente(nome)
                except Exception as e:
                    QMessageBox.critical(self, "Erro", f"Erro ao remover gerente do banco:\n{e}")
                    return
                self.cb_ger.removeItem(idx)